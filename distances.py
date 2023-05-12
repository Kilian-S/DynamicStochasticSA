import googlemaps
import openpyxl
from openpyxl import Workbook, load_workbook


gmaps = googlemaps.Client(key='AIzaSyDDgmthv161tSfmFVmglxlEuJsxn7WnH9A')


class Location:
    def __init__(self, name, latitude, longitude):
        self.name = name
        self.latitude = latitude
        self.longitude = longitude


def create_locations(f):
    workbook = load_workbook(filename=f)
    sheet = workbook.active
    locations = []

    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        latitude = sheet.cell(row=row, column=9).value
        longitude = sheet.cell(row=row, column=10).value

        # Create a Location object and append it to the list
        location = Location(name, latitude, longitude)
        locations.append(location)

    return locations


def create_distance_matrix(locations, max_elements=10):
    n = len(locations)
    distance_matrix = [[0] * n for _ in range(n)]
    coordinates = [(location.latitude, location.longitude) for location in locations]

    for i in range(0, n, max_elements):
        for j in range(0, n, max_elements):
            origins = coordinates[i:min(i + max_elements, n)]
            destinations = coordinates[j:min(j + max_elements, n)]

            # Call the Distance Matrix API with the current set of origins and destinations
            distance_matrix_response = gmaps.distance_matrix(origins, destinations, mode="driving")

            for k, row in enumerate(distance_matrix_response["rows"]):
                for l, element in enumerate(row["elements"]):
                    distance_matrix[i + k][j + l] = element["distance"]["value"]

    return distance_matrix


def write_distance_matrix_to_excel(distance_matrix, file_name, sheet_name):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_name)

    # Check if the sheet_name exists in the workbook, if not, create it
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    # Select the sheet
    sheet = workbook[sheet_name]

    # Write the distance matrix to the sheet
    for i, row in enumerate(distance_matrix, start=2):  # Start at row 2
        for j, distance in enumerate(row, start=2):  # Start at column B
            sheet.cell(row=i, column=j, value=distance)

    # Save the workbook
    workbook.save(file_name)


# locations = create_locations("distances.xlsx")
# matrix = create_distance_matrix(locations)
# write_distance_matrix_to_excel(matrix, "distances.xlsx", "Sheet2")


print("HI")




