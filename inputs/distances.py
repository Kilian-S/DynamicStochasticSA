import googlemaps
import numpy as np
import openpyxl
from openpyxl import load_workbook

GMAPS = googlemaps.Client(key='AIzaSyDDgmthv161tSfmFVmglxlEuJsxn7WnH9A')


class Location:
    def __init__(self, name, latitude, longitude):
        self.name = name
        self.latitude = latitude
        self.longitude = longitude

    def __repr__(self):
        return f"({self.latitude}, {self.longitude}: {self.name}"


def create_locations(filename: str) -> list[Location]:
    """
    Reads a workbook from the given filename, extracts location information from the active sheet,
    and returns a list of Location objects.

    Args:
        filename (str): The name of the workbook file to load.

    Returns:
        list[Location]: A list of Location objects containing name, latitude, and longitude information.

    """
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    locations = []

    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        latitude = sheet.cell(row=row, column=9).value
        longitude = sheet.cell(row=row, column=10).value

        # Create a Location object and append it to the list
        location = Location(name, latitude, longitude)
        locations.append(location)

    return locations


def create_distance_matrix(locations: list[Location], max_elements: int = 10) -> list[list[int]]:
    """
    Creates a distance matrix based on the given list of locations.

    Args:
        locations (list[Location]): A list of Location objects.
        max_elements (int): Maximum number of elements to process in a single API call. Defaults to 10.

    Returns:
        list[list[int]]: A 2D list representing the distance matrix.

    """
    n = len(locations)
    distance_matrix = [[0] * n for _ in range(n)]
    coordinates = [(location.latitude, location.longitude) for location in locations]

    for i in range(0, n, max_elements):
        for j in range(0, n, max_elements):
            origins = coordinates[i:min(i + max_elements, n)]
            destinations = coordinates[j:min(j + max_elements, n)]

            # Call the Distance Matrix API with the current set of origins and destinations
            distance_matrix_response = GMAPS.distance_matrix(origins, destinations, mode="driving")

            for k, row in enumerate(distance_matrix_response["rows"]):
                for l, element in enumerate(row["elements"]):
                    distance_matrix[i + k][j + l] = element["distance"]["value"]

    return distance_matrix


def write_distance_matrix_to_excel(distance_matrix: list[list[int]], file_name: str, sheet_name: str):
    """
    Writes the distance matrix to an Excel file.

    Args:
        distance_matrix (list[list[int]]): The distance matrix to be written.
        file_name (str): The name of the Excel file.
        sheet_name (str): The name of the sheet to write the distance matrix.

    """
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_name)

    # Check if the sheet_name exists in the workbook, if not, create it
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    # Select the sheet
    sheet = workbook[sheet_name]

    # Write the distance matrix to the sheet
    for i, row in enumerate(distance_matrix, start=2):  # Start at row 2
        for j, distance in enumerate(row, start=2):  # Start at column B
            sheet.cell(row=i, column=j, value=distance)

    # Save the workbook
    workbook.save(file_name)


def make_symmetric(matrix: list[list[int]]) -> list[list[int]]:
    """
    Makes a square matrix symmetric by copying values from the upper triangular part to the lower triangular part.

    Args:
        matrix (list[list[int]]): The input matrix.

    Returns:
        list[list[int]]: The resulting symmetric matrix.

    """
    n = len(matrix)
    for i in range(n):
        for j in range(i + 1, n):
            # Use the value from the upper triangular part to update the lower triangular part
            matrix[j][i] = matrix[i][j]
    return matrix


def create_symmetric_distance_matrix(input_file: str, output_file: str, output_sheet: str):
    """
    Creates a symmetric distance matrix from input locations data and writes it to an Excel file.

    Args:
        input_file (str): The path to the input file containing location data.
        output_file (str): The path to the output Excel file to write the distance matrix.
        output_sheet (str): The name of the sheet in the output file to write the distance matrix.

    """
    locations = create_locations(input_file)
    matrix = create_distance_matrix(locations)
    matrix = make_symmetric(matrix)
    write_distance_matrix_to_excel(matrix, output_file, output_sheet)


def read_in_distance_matrix(input_file: str, input_sheet: str, topleft: str, bottomright: str):
    """
    Reads a distance matrix from an Excel file.

    Args:
        input_file (str): The path to the input file.
        input_sheet (str): The name of the sheet in the input file containing the distance matrix.
        topleft (str): The cell reference of the top-left cell of the distance matrix range.
        bottomright (str): The cell reference of the bottom-right cell of the distance matrix range.

    Returns:
        numpy.ndarray: The distance matrix as a numpy array.

    """
    workbook = load_workbook(filename=input_file, read_only=True)
    sheet = workbook[input_sheet]
    cell_range = sheet[topleft:bottomright]
    distance_matrix = [[cell.value for cell in row] for row in cell_range]
    distance_matrix = np.array(distance_matrix)
    return distance_matrix


def normalise_geo_coordinates(input_file: str, new_origin: tuple):
    """
    Normalizes the geographic coordinates of locations in an input file based on a new origin.

    Args:
        input_file (str): The path to the input file.
        new_origin (tuple): The coordinates (latitude, longitude) of the new origin.

    Returns:
        list[Location]: A list of Location objects with normalized coordinates.

    """
    locations = create_locations(input_file)
    origin_latitude, origin_longitude = new_origin
    for location in locations:
        location.latitude = round(location.latitude - origin_latitude, 6)
        location.longitude = round(location.longitude - origin_longitude, 6)
    return locations
