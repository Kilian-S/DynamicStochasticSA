from openpyxl import load_workbook
import numpy as np


class Node:
    """These are the nodes that are seen in the solving process of the SDCVRP."""
    def __init__(self, id, expected_demand):
        self.id = id  # Node ID
        self.expected_demand = expected_demand  # Demand of the node

    def __repr__(self):
        return f"Node {self.id}, Demand: {self.expected_demand}"


class InputNode(Node):
    """These nodes are the nodes that are entered as initial parameters of the SDCVRP. They correspond to actual physical nodes."""
    def __init__(self, id, expected_demand, actual_demand=None):
        super().__init__(id, expected_demand)
        self.actual_demand = actual_demand if actual_demand is not None else 0

    def __repr__(self):
        return f"Node {self.id}, Demand: {self.expected_demand}, Actual Demand: {self.actual_demand}"


def create_nodes_static(filename: str, sheet_name: str):
    workbook = load_workbook(filename=filename)

    # Use the provided sheet name instead of the active sheet
    sheet = workbook[sheet_name]

    nodes = []

    for row in range(2, sheet.max_row + 1):
        id = row - 2
        expected_demand = sheet.cell(row=row, column=3).value
        actual_demand = expected_demand

        # Create a Location object and append it to the list
        node = InputNode(str(id), expected_demand, actual_demand)
        nodes.append(node)

    return nodes


def create_nodes_cauchy(filename: str, sheet_name: str, gamma: float):
    workbook = load_workbook(filename=filename)

    # Use the provided sheet name instead of the active sheet
    sheet = workbook[sheet_name]

    nodes = []

    for row in range(2, sheet.max_row + 1):
        id = row - 2

        if id == 0:
            nodes.append(InputNode(str(id), 0, 0))
            continue

        expected_demand = sheet.cell(row=row, column=3).value

        # Generate actual_demand from a Cauchy distribution
        actual_demand = np.random.standard_cauchy() * gamma + expected_demand

        # Ensure that actual_demand is non-negative
        actual_demand = max(0, actual_demand)

        # Create a Location object and append it to the list
        node = InputNode(str(id), expected_demand, int(actual_demand))
        nodes.append(node)

    return nodes


def create_nodes_cauchy_dependent_on_expected_demand(filename: str, sheet_name: str, gamma_factor: float):
    workbook = load_workbook(filename=filename)

    # Use the provided sheet name instead of the active sheet
    sheet = workbook[sheet_name]

    nodes = []

    for row in range(2, sheet.max_row + 1):
        id = row - 2

        if id == 0:
            nodes.append(InputNode(str(id), 0, 0))
            continue

        expected_demand = sheet.cell(row=row, column=3).value
        gamma = expected_demand * gamma_factor

        # Generate actual_demand from a Cauchy distribution
        actual_demand = np.random.standard_cauchy() * gamma + expected_demand

        # Ensure that actual_demand is non-negative
        actual_demand = max(0, actual_demand)

        # Create a Location object and append it to the list
        node = InputNode(str(id), expected_demand, int(actual_demand))
        nodes.append(node)

    return nodes





