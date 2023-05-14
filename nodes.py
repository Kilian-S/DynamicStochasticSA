import openpyxl
import os
from openpyxl import load_workbook


class Node:
    def __init__(self, id, demand):
        self.id = id  # Node ID
        self.demand = demand  # Demand of the node

    def __str__(self):
        return f"Node {self.id}, Demand: {self.demand}"


def create_nodes(f):
    workbook = load_workbook(filename=f)
    sheet = workbook.active
    nodes = []

    for row in range(2, sheet.max_row + 1):
        id = row-2
        demand = sheet.cell(row=row, column=3).value

        # Create a Location object and append it to the list
        node = Node(id, demand)
        nodes.append(node)

    return nodes


print(create_nodes("distances.xlsx")[10])
