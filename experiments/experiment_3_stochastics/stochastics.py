import ast
from openpyxl import Workbook
from dynamic_behaviour import dynamic_sa, get_tour_demand
from inputs.distances import read_in_distance_matrix
from inputs.dynamic_nodes_list import DynamicNodeList
from inputs.node import create_nodes_cauchy_dependent_on_expected_demand, InputNode, create_nodes_static, create_nodes_cauchy, Node
from inputs.node_family import NodeFamily
from simulated_annealing import objective
import pandas as pd


def get_omniscient_nodes(stochastic_nodes: list[InputNode], vehicle_capacity: int):
    node_families = [NodeFamily(node, vehicle_capacity) for node in stochastic_nodes]
    dynamic_node_list = DynamicNodeList(node_families, vehicle_capacity)
    for node_family in node_families[1:]:
        node_family.update()
    omniscient_nodes = dynamic_node_list.get_all_nodes()
    return omniscient_nodes


def get_total_oversupply(tours: list[list[str]], nodes: list[Node], vehicle_capacity: int, tours_set: set[str], nodes_set: set[str]) -> int:
    # Create a dictionary for faster node access
    nodes_dict = {node.id: node for node in nodes}

    total_oversupply = 0

    for tour in tours:
        tour_demand = 0

        for node_id in tour:
            node = nodes_dict.get(node_id)
            if node is not None:
                tour_demand += node.expected_demand

        tour_oversupply = max(0, vehicle_capacity - tour_demand)
        total_oversupply += tour_oversupply

    # Handle nodes not considered by the exact solution
    if tours_set != nodes_set:
        difference_set = nodes_set - tours_set

        for node_id in difference_set:
            total_oversupply += vehicle_capacity - nodes_dict.get(node_id).expected_demand

    return total_oversupply


def get_total_undersupply(tours: list[list[str]], nodes: list[Node], vehicle_capacity: int, tours_set: set[str], nodes_set: set[str]) -> tuple:
    # Create a dictionary for faster node access
    nodes_dict = {node.id: node for node in nodes}
    total_undersupply = 0
    undersupplied_tours_count = 0

    for tour in tours:
        tour_demand = 0

        for node_id in tour:
            node = nodes_dict.get(node_id)
            if node is not None:
                tour_demand += node.expected_demand

        if tour_demand > vehicle_capacity:
            total_undersupply += tour_demand - vehicle_capacity
            undersupplied_tours_count += 1

    # Handle nodes not considered by the exact solution
    if tours_set != nodes_set:
        difference_set = nodes_set - tours_set

        for node_id in difference_set:
            total_undersupply += nodes_dict.get(node_id).expected_demand
            undersupplied_tours_count += 1

    return total_undersupply, undersupplied_tours_count


def get_service_level(tours: list[list[str]], nodes: list[Node], vehicle_capacity: int, tours_set: set[str], nodes_set: set[str]) -> float:
    total_network_demand = 0
    served_demand = 0
    nodes_dict = {node.id: node for node in nodes}

    for tour in tours:
        tour_demand = 0  # Track total demand for this tour
        remaining_capacity = vehicle_capacity  # Reset the remaining vehicle capacity for each tour

        for node_id in tour:
            node = nodes_dict.get(node_id)
            if node is not None:
                node_demand = node.expected_demand
                tour_demand += node_demand

                # Count the node's demand towards served demand only if there's still capacity left
                if remaining_capacity > 0:
                    demand_served_at_this_node = min(node_demand, remaining_capacity)
                    served_demand += demand_served_at_this_node
                    remaining_capacity -= demand_served_at_this_node

        total_network_demand += tour_demand

    # Add demand of nodes not considered in the static solution
    if tours_set != nodes_set:
        difference_set = nodes_set - tours_set
        for node_id in difference_set:
            node = nodes_dict.get(node_id)
            if node is not None:
                total_network_demand += node.expected_demand

    # Avoid division by zero error
    if total_network_demand == 0:
        return 0.0

    return served_demand / total_network_demand


def stochastic_nodes_to_excel(n: int, gamma_values: list[float], filename='../../inputs/distances.xlsx', sheetname='Sheet1'):
    # Assuming create_nodes_static and create_nodes_cauchy are functions that return lists of InputNode objects
    static_nodes = create_nodes_static(filename, sheetname)

    workbook = Workbook()

    # Remove default sheet if it exists
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Writing the static nodes to the Excel file
    static_demand = [node.actual_demand for node in static_nodes]

    # Generating n sets of nodes with Cauchy distributed demand for each gamma value
    for gamma in gamma_values:
        # Create a new sheet for this gamma value
        sheet = workbook.create_sheet(title=f"Gamma_{gamma}")
        sheet.append(static_demand)

        for _ in range(n):
            stochastic_nodes = create_nodes_cauchy(filename, sheetname, gamma)
            stochastic_demand = [node.actual_demand for node in stochastic_nodes]
            sheet.append(stochastic_demand)

    # Saving the Excel file
    #workbook.save('comparison.xlsx')


def stochastic_nodes_to_excel_dependent_on_expected_demand(n: int, gamma_factors: list[float], filename='../../inputs/distances.xlsx', sheetname='Sheet1'):
    # Assuming create_nodes_static and create_nodes_cauchy are functions that return lists of InputNode objects
    static_nodes = create_nodes_static(filename, sheetname)

    workbook = Workbook()

    # Remove default sheet if it exists
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Writing the static nodes to the Excel file
    static_demand = [node.actual_demand for node in static_nodes]

    # Generating n sets of nodes with Cauchy distributed demand for each gamma value
    for gamma_factor in gamma_factors:
        # Create a new sheet for this gamma factor value
        sheet = workbook.create_sheet(title=f"Gamma_Factor_{gamma_factor}")
        sheet.append(static_demand)

        for _ in range(n):
            stochastic_nodes = create_nodes_cauchy_dependent_on_expected_demand(filename, sheetname, gamma_factor)
            stochastic_demand = [node.actual_demand for node in stochastic_nodes]
            sheet.append(stochastic_demand)

    # Saving the Excel file
    #workbook.save('comparison_normalised.xlsx')


def stochastic_nodes_experiment(trials: int, gamma_values: list[float], output_filename: str):
    initial_temperature = 100
    iterations = 1000
    utilisation_target = 0.9
    vehicle_capacity = 2000
    distance_matrix = read_in_distance_matrix('../../inputs/distances.xlsx', 'Distance matrix (districts)', 'B2', 'AX50')

    # Exact solution
    df = pd.read_excel(r'C:\Kilian\TUM\TUM\Bachelor Thesis\Code\simulated annealing\experiments\experiment_2_perfect_information\results_static.xlsx')
    df['tours'] = df['tours'].apply(ast.literal_eval)
    exact_tours = df['tours'].iloc[0]
    exact_tours_set = set(string for tour in exact_tours for string in tour)

    data = {'gamma_value': [], 'trial': [], 'demands': [], 'objective_value': [], 'tours': [], 'execution_time': [], 'service_level_sa': [], 'total_oversupply_sa': [],
            'total_undersupply_sa': [], 'total_undersupply_sa_count': [], 'service_level_exact': [], 'total_oversupply_exact': [], 'total_undersupply_exact': [],
            'total_undersupply_exact_count': []}

    # Generating n sets of nodes with Cauchy distributed demand for each gamma value
    for gamma_value in gamma_values:
        for trial in range(trials):
            stochastic_nodes = create_nodes_cauchy_dependent_on_expected_demand('../../inputs/distances.xlsx', 'Sheet1', gamma_value)
            stochastic_demand = [node.actual_demand for node in stochastic_nodes]
            current_tours_value, current_tours, execution_time, omniscient_nodes = dynamic_sa(stochastic_nodes, distance_matrix, objective, initial_temperature, iterations,
                                                                                              vehicle_capacity, utilisation_target)

            current_tours_set = set(string for tour in current_tours for string in tour)
            omniscient_nodes_set = set(node.id for node in omniscient_nodes)

            service_level_sa = get_service_level(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            total_oversupply_sa = get_total_oversupply(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            total_undersupply_sa, total_undersupply_sa_count = get_total_undersupply(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            service_level_exact = get_service_level(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)
            total_oversupply_exact = get_total_oversupply(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)
            total_undersupply_exact, total_undersupply_exact_count = get_total_undersupply(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)

            # Append data to the dictionary
            data['gamma_factor'].append(gamma_value)
            data['trial'].append(trial + 1)
            data['demands'].append(stochastic_demand)
            data['objective_value'].append(current_tours_value)
            data['tours'].append(str(current_tours))
            data['execution_time'].append(execution_time)
            data['service_level_sa'].append(service_level_sa)
            data['total_oversupply_sa'].append(total_oversupply_sa)
            data['total_undersupply_sa'].append(total_undersupply_sa)
            data['total_undersupply_sa_count'].append(total_undersupply_sa_count)
            data['service_level_exact'].append(service_level_exact)
            data['total_oversupply_exact'].append(total_oversupply_exact)
            data['total_undersupply_exact'].append(total_undersupply_exact)
            data['total_undersupply_exact_count'].append(total_undersupply_exact_count)

        # Create a DataFrame from the dictionary
        df = pd.DataFrame(data)

        # Write the DataFrame to an Excel file
        df.to_excel(output_filename, index=False)


def stochastic_nodes_dependent_on_expected_demand_experiment(trials: int, gamma_factors: list[float], output_filename: str):
    initial_temperature = 100
    iterations = 1000
    utilisation_target = 0.9
    vehicle_capacity = 2000
    distance_matrix = read_in_distance_matrix('../../inputs/distances.xlsx', 'Distance matrix (districts)', 'B2', 'AX50')

    # Exact solution
    df = pd.read_excel(r'C:\Kilian\TUM\TUM\Bachelor Thesis\Code\simulated annealing\experiments\experiment_2_perfect_information\results_static.xlsx')
    df['tours'] = df['tours'].apply(ast.literal_eval)
    exact_tours = df['tours'].iloc[0]
    exact_tours_set = set(string for tour in exact_tours for string in tour)

    # Create a dictionary to store data for the DataFrame
    data = {'gamma_factor': [], 'trial': [], 'demands': [], 'objective_value': [], 'tours': [], 'execution_time': [], 'service_level_sa': [], 'total_oversupply_sa': [],
            'total_undersupply_sa': [], 'total_undersupply_sa_count': [], 'service_level_exact': [], 'total_oversupply_exact': [], 'total_undersupply_exact': [],
            'total_undersupply_exact_count': []}

    # Generating n sets of nodes with Cauchy distributed demand for each gamma value
    for gamma_factor in gamma_factors:
        for trial in range(trials):
            stochastic_nodes = create_nodes_cauchy_dependent_on_expected_demand('../../inputs/distances.xlsx', 'Sheet1', gamma_factor)
            stochastic_demand = [node.actual_demand for node in stochastic_nodes]
            current_tours_value, current_tours, execution_time, omniscient_nodes = dynamic_sa(stochastic_nodes, distance_matrix, objective, initial_temperature, iterations,
                                                                                              vehicle_capacity, utilisation_target)

            current_tours_set = set(string for tour in current_tours for string in tour)
            omniscient_nodes_set = set(node.id for node in omniscient_nodes)

            service_level_sa = get_service_level(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            total_oversupply_sa = get_total_oversupply(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            total_undersupply_sa, total_undersupply_sa_count = get_total_undersupply(current_tours, omniscient_nodes, vehicle_capacity, current_tours_set, omniscient_nodes_set)
            service_level_exact = get_service_level(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)
            total_oversupply_exact = get_total_oversupply(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)
            total_undersupply_exact, total_undersupply_exact_count = get_total_undersupply(exact_tours, omniscient_nodes, vehicle_capacity, exact_tours_set, omniscient_nodes_set)

            # Append data to the dictionary
            data['gamma_factor'].append(gamma_factor)
            data['trial'].append(trial + 1)
            data['demands'].append(stochastic_demand)
            data['objective_value'].append(current_tours_value)
            data['tours'].append(str(current_tours))
            data['execution_time'].append(execution_time)
            data['service_level_sa'].append(service_level_sa)
            data['total_oversupply_sa'].append(total_oversupply_sa)
            data['total_undersupply_sa'].append(total_undersupply_sa)
            data['total_undersupply_sa_count'].append(total_undersupply_sa_count)
            data['service_level_exact'].append(service_level_exact)
            data['total_oversupply_exact'].append(total_oversupply_exact)
            data['total_undersupply_exact'].append(total_undersupply_exact)
            data['total_undersupply_exact_count'].append(total_undersupply_exact_count)

            # Create a DataFrame from the dictionary
            df = pd.DataFrame(data)

            # Write the DataFrame to an Excel file
            df.to_excel(output_filename, index=False)


stochastic_nodes_dependent_on_expected_demand_experiment(35, [0.15], 'results_stochastics_cauchy_gamma_factors_app.xlsx')













